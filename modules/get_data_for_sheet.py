#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
模块名称：get_data_for_sheet.py
----------------------------------------
功能：根据 report_id，从数据库中获取数据，填充 Excel 各 Sheet 表格
调用方式：由服务调度程序 server_detection.py 触发
"""

import os
import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine
import logging

# 设置日志
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

# 数据库路径（可改为配置项）
DB_PATH = "inspection.db"

# Excel 模板路径（可改为配置项）
EXCEL_TEMPLATE = "../data/巡检报告数据集(1.0).xlsx"

# 输出目录
OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

def run_data_fill_pipeline(report_id: str) -> None:
    """
    主流程：从数据库中查询业务数据，填入 Excel 模板中各个 sheet。
    参数：
        report_id : str，巡检报告唯一编号，用于数据筛选、输出命名。
    返回：
        无，直接保存 Excel 文件（供后续模块使用）
    """
    print(f"[模拟] ✅ 模拟连接数据库，获取数据填写 excel 文件 sheet 表格成功")
    return

    try:
        # 加载 Excel 模板
        excel_path = os.path.join(OUTPUT_DIR, f"{report_id}.xlsx")
        wb = load_workbook(EXCEL_TEMPLATE)

        # 连接数据库
        engine = create_engine(f"sqlite:///{DB_PATH}")
        conn = engine.connect()

        # 示例填充：将每个 sheet 填充对应表数据
        sheet_table_map = {
            "表1": "sheet1_data",
            "表2": "sheet2_data",
            # 可继续扩展
        }

        for sheet_name, table_name in sheet_table_map.items():
            df = pd.read_sql(f"SELECT * FROM {table_name} WHERE report_id = ?", conn, params=[report_id])
            ws = wb[sheet_name]
            for r_idx, row in enumerate(df.itertuples(index=False), start=2):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            log.info(f"✅ 已填充 sheet: {sheet_name}")

        # 保存新文件
        wb.save(excel_path)
        log.info(f"✅ 已保存 Excel: {excel_path}")

    except Exception as e:
        log.error(f"❌ 数据填充失败: {e}")
