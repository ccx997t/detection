#!/usr/bin/env python3                     # 指定解释器为 Python3，可在命令行直接执行该脚本
# -*- coding: utf-8 -*-                    # 指定文件编码为 UTF-8，确保中文注释和日志正常显示
"""
巡检报告模板Excel转换jpg文件模块（excel_to_images.py）
Excel → JPG 图像生成模块（统一 LibreOffice 渲染 + 日志版）
------------------------------------------------------------
功能：
    将 Excel 文件通过 LibreOffice 无头模式渲染为高保真 PDF，
    再使用 pdf2image 将 PDF 转为 JPG。
依赖：
    libreoffice、poppler-utils、pandas、pdf2image、Pillow
"""

# ============================================================
# 导入模块
# ============================================================
import os                                  # 提供文件和路径操作函数
import sys                                 # 提供系统级访问，如路径与退出
import configparser                        # 配置解释器。
import subprocess                          # 用于执行外部命令（调用 LibreOffice）
from pdf2image import convert_from_path    # 将 PDF 转换为 JPG 的核心函数
from PIL import Image, ImageChops          # 处理图像（裁剪空白边）所需模块
from typing import Dict, List              # 类型标注，用于提高代码可读性
from openpyxl import load_workbook          # 替代 pandas 用于读取 sheet
import tempfile
import shutil
from pathlib import Path

# ============================================================
# 修正项目模块搜索路径
# ============================================================
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # 获取项目根目录
if PROJECT_ROOT not in sys.path:               # 若项目根目录未加入 Python 模块搜索路径
    sys.path.append(PROJECT_ROOT)              # 动态添加，以便能导入项目自定义模块
print(f">> PROJECT_ROOT = {PROJECT_ROOT},  __file__ = {__file__}")  # 打印当前项目根路径

# ============================================================
# 项目模块 util
# ============================================================
try:
    from modules import util as _ut
except Exception as e:
    _ut = None
    print(f"⚠️  未找到 util 模块：{e}")

# 实例化日志类
log = _ut.Logger()

# 全局参数
EXCEL_PATH = ""
PDFS_DIR = ""
IMAGES_DIR = ""
OUTPUT_DIR = ""
PAGE_SIZE = 0
ORIENTATION = ""
DPI = 0

def crop_whitespace(image_path: str):
    """裁剪 JPG 图像四周的空白边。"""
    img = Image.open(image_path)              # 打开指定的图像文件
    bg = Image.new(img.mode, img.size, img.getpixel((0, 0)))  # 生成一张背景色相同的空图
    diff = ImageChops.difference(img, bg)     # 计算原图与背景的差异区域
    bbox = diff.getbbox()                     # 获取有效内容的边界框
    if bbox:                                  # 如果存在非空白区域
        cropped = img.crop(bbox)              # 裁剪图像到内容区域
        cropped.save(image_path)              # 覆盖保存原文件
        log.info(f"已裁剪白边：{image_path}")  # 输出日志提示裁剪完成

def load_sheet_names() -> List[str]:
    """ 使用 openpyxl 加载 Excel 文件，提取所有工作表名称。"""
    # 以只读模式打开 Excel 文件，提高加载效率
    wb = load_workbook(EXCEL_PATH, read_only=True)
    # 获取当前 Excel 中的所有工作表名称
    names = wb.sheetnames 
    # 关闭文件，释放资源
    wb.close()
    # 返回 sheet 名称列表
    return names

def adjust_excel() -> str:
    """ Excel 页面设置预处理模块,将 Excel 每个 sheet 设置为“单页模式”，供 LibreOffice 转 PDF 时使用。
    """
    log.info(f"🔧 开始调整 Excel 打印配置为单页模式：{EXCEL_PATH}")
    try:
        # PDFS_DIR目录默认是"tmp/pdfs", 调整后的临时文件保存在"tmp/"目录下。
        tmp_dir = os.path.dirname(os.path.abspath(PDFS_DIR))

        os.makedirs(tmp_dir, exist_ok=True)

        # 删除临时目录下的文件。
        _ut.remove_path_recursio_files_func(tmp_dir)        

        # 1️ 创建临时目录并复制原始 Excel 文件
        adjusted_path = _ut.gen_target_file_name_func(EXCEL_PATH, tmp_dir, "临时")
        # 将原本复制一个副本。
        shutil.copy2(EXCEL_PATH, adjusted_path)
        log.info(f"📁 创建临时副本：{adjusted_path}")
 
        # 2️ 使用 openpyxl 加载副本
        wb = load_workbook(adjusted_path)
        modified_count = 0
        total_sheets = len(wb.worksheets)
        log.info(f"📄 加载副本成功，共包含 {total_sheets} 个工作表")
  
        # 3️ 遍历每个 sheet，应用打印设置
        for idx, sheet in enumerate(wb.worksheets, start=1):
            log.info(f"🔍 正在处理第 {idx} 个 sheet：{sheet.title}")
            # 设置打印缩放参数，确保整个sheet压缩为单页显示
            ps = sheet.page_setup
            ps.fitToWidth = 1                 # 一页宽度内显示全部列
            ps.fitToHeight = 1                # 一页高度内显示全部行
            ps.scale = None                   # 禁止自定义比例，避免与fitToPage冲突
            ps.paperSize = PAGE_SIZE          # 纸张类型编号（A3或A4）
            ps.orientation = ORIENTATION      # 纵向打印
            # 启用“适应单页打印”模式
            sheet.sheet_properties.pageSetUpPr.fitToPage = True
            # 自动计算并设置打印区域，确保导出时包含所有单元格
            sheet.print_area = sheet.calculate_dimension()
            # 设置打印输出居中显示（水平+垂直）
            sheet.print_options.horizontalCentered = True
            sheet.print_options.verticalCentered = True

        # 4️ 保存副本
        wb.save(adjusted_path)
        wb.close()
        log.info(f"💾 保存完成，已修改 {modified_count} 个 sheet")

        # 5️ 返回副本路径
        log.info(f"✅ Excel 页面调整完成，输出路径：{adjusted_path}")
        #print(f">> adjusted_path = {adjusted_path}")
        return adjusted_path

    except Exception as e:
        log.error(f"❌ 出现错误：{str(e)}")
        return adjusted_path

def excel_to_libreoffice_pdf() -> str:
    """调用 LibreOffice 将 Excel 转换为 PDF。"""
    # 调整输入的 Excel 文件为单页 sheet Excel文件。
    adjusted_excel_path = adjust_excel()
    # 将 Excel 渲染为PDF。
    log.info("使用 soffice --headless 渲染 Excel → PDF ...")
    try:
        # 组装 soffice 命令
        cmd = [
            "soffice",
            "--headless",                           # 无界面模式
            "--convert-to", "pdf",                  # 输出格式 PDF
            "--outdir", PDFS_DIR,       # 输出目录
            adjusted_excel_path         # 输入文件路径
        ]
        # 显示执行命令信息。
        log.info(f"执行命令：{' '.join(cmd)}")
        # 执行命令行调用
        result = subprocess.run(cmd, capture_output=True, text=True)
        # 检查返回状态
        if result.returncode != 0:
            log.error(f"LibreOffice 转换失败：{result.stderr.strip()}")
            raise RuntimeError(f"LibreOffice 转换失败：{result.stderr.strip()}")
        # 为 pdf_path 赋值。
        pdf_path = os.path.join(PDFS_DIR, Path(adjusted_excel_path).stem + ".pdf")
    except Exception as e:
        log.error(f"Excel → PDF 渲染异常：{e}")
        raise
    log.info(f"✅ 已生成 PDF：{pdf_path}")
    # 返回生成的 PDF 路径
    return pdf_path 

def pdf_to_jpgs(pdf_path: str, sheet_names: List[str]) -> Dict[str, str]:
    """将 PDF 多页转换为 JPG 并与 sheet 对齐命名。"""
    # 确保 JPG 输出目录存在
    os.makedirs(IMAGES_DIR, exist_ok=True) 
    # 输出开始转换日志
    log.info("开始 PDF → JPG 拆分 ...") 
    # 调用 pdf2image 将 PDF 每页转为图像对象
    images = convert_from_path(pdf_path, DPI, fmt="jpeg") 
    # 获取 PDF 页数与 Excel 工作表数量
    num_pages, num_sheets = len(images), len(sheet_names) 
    # 输出对比信息
    log.info(f"PDF 页数：{num_pages}，Excel 工作表数：{num_sheets}") 
    # 初始化映射字典：sheet_name → JPG 文件路径
    mapping: Dict[str, str] = {}
    # 遍历每一页 PDF 图像
    for i, img in enumerate(images):          
        # 若页数多于 sheet，用 PageX 命名
        name = sheet_names[i] if i < num_sheets else f"Page{i+1}" 
        # 生成 JPG 输出路径
        jpg_path = os.path.join(IMAGES_DIR, f"{name}.jpg")         
        # 保存当前页为 JPG 文件
        img.save(jpg_path, "JPEG")          
        # 调用函数裁剪白边
        crop_whitespace(jpg_path)            
        # 记录映射关系
        mapping[name] = jpg_path             
        # 输出日志
        log.info(f"生成 JPG：{jpg_path}")  

    # 若 PDF 页数少于 sheet 数，说明部分表未匹配
    if num_pages < num_sheets:
        # 输出警告信息
        log.warn(f"以下 sheet 未匹配到页面：{sheet_names[num_pages:]}")
    # 返回 sheet → JPG 的映射关系
    return mapping                            

def excel_to_jpgs() -> Dict[str, str]:
    """主函数：Excel → PDF → JPG """
    if not os.path.exists(EXCEL_PATH):               # 若输入 Excel 文件不存在
        log.error(f"Excel 文件不存在：{EXCEL_PATH}")  # 输出错误
        raise FileNotFoundError(EXCEL_PATH)          # 抛出异常终止程序

    # 获取所有工作表名
    sheet_names = load_sheet_names() 
    # excel → PDF
    pdf_path = excel_to_libreoffice_pdf()
    # PDF → JPG 拆页转换
    mapping = pdf_to_jpgs(pdf_path, sheet_names) 
    # 输出完成信息
    log.info("🎯 所有 JPG 文件已生成。")
    # 返回转换结果字典
    return mapping

def run(config: configparser.ConfigParser):
    """外部调用接口。"""    
    # 提取配置文件参数项
    global EXCEL_PATH, PDFS_DIR, IMAGES_DIR, OUTPUT_DIR, PAGE_SIZE, ORIENTATION, DPI
    EXCEL_PATH = config.get("Path", "input_path")
    PDFS_DIR = config.get("Path", "pdfs_dir")
    IMAGES_DIR = config.get("Path", "images_dir")
    OUTPUT_DIR = config.get("Path", "output_dir")
    PAGE_SIZE = config.getint("PageConf", "page_size")
    ORIENTATION = config.get("PageConf", "orientation")
    DPI = config.getint("PageConf", "dpi")

    log.info("run() 启动 Excel → JPG 转换流程")     # 输出流程开始日志
    mapping = excel_to_jpgs()                      # 调用主函数执行转换
    log.info("=== 输出文件映射 ===")                # 输出结果映射表头
    for k, v in mapping.items():                   # 遍历每个 sheet 对应的 JPG 文件
        log.info(f"{k} -> {v}")                    # 输出映射关系日志
    log.info(f"✅ 输出目录：{IMAGES_DIR}")          # 输出最终目录路径

# ============================================================
# main 程序入口
# ============================================================
if __name__ == "__main__":                           # 若脚本以主程序方式运行
    """命令行入口函数（教学示例）。"""
    #EXCEL_PATH = "../data/销售统计表.xlsx"           # 设置输入 Excel 文件路径
    EXCEL_PATH = "../data/巡检报告数据集(1.0).xlsx"   # 设置输入 Excel 文件路径    
    PDFS_DIR = "../tmp/pdfs/"                        # 设置 PDF 输出目录
    IMAGES_DIR = "../tmp/images/"                    # 设置 JPG 输出目录
    PAGE_SIZE = 8
    ORIENTATION = "portrait"
    DPI = 300                                        # 设置转换分辨率（打印级清晰度）
    log.info("run() 启动 Excel → JPG 转换流程")       # 输出流程开始日志
    mapping = excel_to_jpgs()                        # 调用主函数执行转换
    log.info("=== 输出文件映射 ===")                  # 输出结果映射表头
    for k, v in mapping.items():                     # 遍历每个 sheet 对应的 JPG 文件
        log.info(f"{k} -> {v}")                      # 输出映射关系日志
    log.info(f"✅ 输出目录：{IMAGES_DIR}")           # 输出最终目录路径
